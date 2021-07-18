import pandas as pd
from openpyxl import load_workbook
from collections import Counter
df = pd.read_excel('Frequency of Purchase Analysis Data Question (2).xlsx', sheet_name='Data')
def unique(list1):  
    unique_list = []
    for x in list1:
        if x not in unique_list:
            unique_list.append(x)
    return unique_list

#replacing underscore by space
a = df.columns = [c.replace(' ', '_') for c in df.columns]
#taking outlets id
list_outlets = df['Outlet_ID'].tolist()
list_unique = unique(list_outlets)

#counting the occurance of outlet_id
occ_val=[]
for i in list_unique:
    val = df[df['Outlet_ID'] == i].shape[0]
    occ_val.append(val)
tot_sales=0.0
tot_sales_list =[]
#calculating total values of sales
df1 = df.set_index("Outlet_ID")
for i in list_unique:
  new_df1 = df1.loc[[i]]
  tot_sales = new_df1['Sales_Value'].sum()
  tot_sales_list.append(tot_sales)

final_count=[]
for i in range(1,10):
  val = occ_val.count(i)
  final_count.append(val)
print(final_count)
df_new = pd.DataFrame({'Number of outlets': final_count})
wb = load_workbook('output.xlsx')
ws = wb['Sheet1']
for index, row in df_new.iterrows():
    cell = 'B%d'  % (index + 2)
    ws[cell] = row[0]
wb.save('output.xlsx')
